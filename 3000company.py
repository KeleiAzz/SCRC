__author__ = 'keleigong'

import openpyxl as ox

companyList = ox.Workbook()
list = companyList.create_sheet(0,'sheet1')

f = open("/Users/keleigong/Dropbox/Python/Project-SCRC/3000.txt")
line = f.readline()
temp = line.split('\r')
print temp
Row = 1
for i in temp:

    name = ' '.join(i.split()[0:-1])
    ticker = i.split()[-1]
    list.cell(row = Row, column = 1).value = name
    list.cell(row = Row, column = 2).value = ticker
    print name + '---' + ticker
    Row += 1

companyList.save('3000company.xlsx')