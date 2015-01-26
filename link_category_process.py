__author__ = 'keleigong'

#input file link_category_temp, add ID to each company, convert category to category_ID

import openpyxl as ox

source = ox.load_workbook("/Users/keleigong/Dropbox/Python/Project-SCRC/link_category_temp.xlsx")
link = source.get_sheet_by_name('link_category_temp')['A1':'D7275']
companyList = source.get_sheet_by_name('sheet')['A1':'B1246']
link_convert = source.create_sheet(2, 'sheet1')

company = {}

for rows in companyList:
    # print rows[1].value
    temp = rows[1].value.encode('utf8').replace(',', ' ').replace("'", "").lower().split()
    # print temp
    if temp[-1] == 'corp' or temp[-1] == 'corp.' or temp[-1] == 'inc' or temp[-1] == 'inc.' or temp[-1] == 'company' \
            or temp[-1] == 'ltd' or temp[-1] == 'ltd.':
        if company.has_key(' '.join(temp[0:-1])):
            print temp
        company[' '.join(temp[0:-1])] = rows[0].value
    else:
        if company.has_key(' '.join(temp)):
            print temp
        company[' '.join(temp)] = rows[0].value
        # company[rows[1].value.encode('utf8')] = rows[0]

row_pointer = 1

category_id = {'SM': 1, 'SS': 2, 'CM': 3, 'SRM': 4, 'LHR': 5, 'SUS': 6}

for rows in link:
    temp = rows[0].value.encode('utf8').replace(',', ' ').replace("'", "").lower().split()
    if temp[-1] == 'corp' or temp[-1] == 'corp.' or temp[-1] == 'inc' or temp[-1] == 'inc.' or temp[-1] == 'company' \
            or temp[-1] == 'ltd' or temp[-1] == 'ltd.':
        companyName = ' '.join(temp[0:-1])
    else:
        companyName = ' '.join(temp)
    # temp = str(temp).replace('\xe2\x80\x99',"")
    companyName = str(companyName).replace('\xe2\x80\x99', "")
    if company.has_key(companyName):
        cid = company[companyName]
        category_set = rows[1].value.encode('utf8').split(',')
        for i in category_set:
            link_convert.cell(row=row_pointer, column=1).value = cid
            link_convert.cell(row=row_pointer, column=2).value = companyName
            link_convert.cell(row=row_pointer, column=3).value = category_id[i]
            link_convert.cell(row=row_pointer, column=4).value = rows[2].value
            link_convert.cell(row=row_pointer, column=5).value = rows[3].value
            row_pointer += 1
    else:
        print companyName


source.save('link_category_converted')
