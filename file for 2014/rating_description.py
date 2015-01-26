__author__ = 'keleigong'

import openpyxl as ox

wb = ox.Workbook()
target = wb.create_sheet(0, 'discription')

source = ox.load_workbook('/Users/keleigong/Dropbox/Python/Project-SCRC/rating_description.xlsx')
source_date = source.get_sheet_by_name('sheet1')

target_row = 1
count = 1
for row_num in range(2,source_date.max_row+1,1):
    primary_key = str(source_date.cell(row = row_num, column=4).value.year) + '-' + str(source_date.cell(row=row_num,column = 3).value)+'-'+str(source_date.cell(row=row_num,column = 2).value) +\
        '-'+str(count)
    target.cell(row = target_row, column = 1).value = primary_key
    target.cell(row = target_row, column = 2).value = source_date.cell(row=row_num,column = 2).value
    target.cell(row = target_row, column = 3).value = source_date.cell(row=row_num,column = 3).value
    target.cell(row = target_row, column = 4).value = source_date.cell(row=row_num,column = 4).value
    target.cell(row = target_row, column = 5).value = source_date.cell(row=row_num,column = 5).value
    target.cell(row = target_row, column = 6).value = source_date.cell(row=row_num,column = 6).value
    if(source_date.cell(row=row_num,column=2).value == source_date.cell(row=row_num+1,column=2).value):
        count += 1
    else:
        count = 1
    target_row += 1

wb.save('rating_description_output.xlsx')