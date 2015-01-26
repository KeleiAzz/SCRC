__author__ = 'keleigong'


import openpyxl as ox

#-------create new workbook and sheet----
stock_prices = ox.Workbook()
prices = stock_prices.create_sheet(0,'prices')


#-------load the exist data-----
Company = ox.load_workbook('Company_list.xlsx')
original_prices = Company.get_sheet_by_name('Stock prices')
Company_f = Company.get_sheet_by_name('Financial details')


Company_ID_Bticker = {} #record the company ID and B_ticker

#id = 1

for i in range(3,len(Company_f.rows)+1,1):
    Bticker = str(Company_f.cell(row = i, column = 11).value)
    Company_ID_Bticker["".join(Bticker.split()).lower()] = int(Company_f.cell(row = i, column = 1).value)
    #id += 1
R = 1

for i in range(len(Company_ID_Bticker)):
    Company_Bticker = str(original_prices.cell(row = 4*i+1, column = 2).value)
    find_key = "".join(Company_Bticker.split()).lower()
    if find_key in Company_ID_Bticker.keys():
        original_prices.cell(row = 4*i+1, column = 1).value = Company_ID_Bticker[find_key]
        j = 3
        while original_prices.cell(row = 4*i+2, column = j).value != None and j <= 87:
            prices.cell(row = R, column = 1).value = Company_ID_Bticker[find_key]
            prices.cell(row = R, column = 2).value = Company_Bticker
            prices.cell(row = R, column = 3).value = original_prices.cell(row = 4*i+2, column = j).value
            prices.cell(row = R, column = 4).value = original_prices.cell(row = 4*i+3, column = j).value
            R += 1
            j += 1
    else:
        print Company_Bticker



#New_book = ox.Workbook()

#NB = New_book.create_sheet(0)
#NB.title = 'Price'
#NB = original_prices

#New_book.save('New book.xlsx')

#original_prices
stock_prices.save('try.xlsx')

