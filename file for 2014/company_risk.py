__author__ = 'keleigong'

import openpyxl as ox

#-------create new workbook and sheet----
#stock_prices = ox.Workbook()
#prices = stock_prices.create_sheet(0,'prices')


#-------load the exist data-----
#Company = ox.load_workbook('/Users/keleigong/Documents/Python/Project-SCRC/Company_list.xlsx')
#original_prices = Company.get_sheet_by_name('Stock prices')
#Company_f = Company.get_sheet_by_name('Financial details')
line =[]
myfile = open('CompanyList.txt')
for lines in myfile:
    line.append(lines[0:-1])

myfile.close()

