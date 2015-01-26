__author__ = 'keleigong'


import openpyxl as ox

#-------create new workbook and sheet----
#stock_prices = ox.Workbook()
#prices = stock_prices.create_sheet(0,'prices')


#-------load the exist data-----
Company = ox.load_workbook('/Users/keleigong/Documents/Python/Project-SCRC/Company_list.xlsx')
#original_prices = Company.get_sheet_by_name('Stock prices')
Company_f = Company.get_sheet_by_name('Financial details')

company_ID_List = []

row_num = 3

while Company_f.cell(row = row_num,column = 1).value != None:
    company_ID_List.append(str(Company_f.cell(row = row_num,column = 1).value)+'@'+\
                            Company_f.cell(row = row_num,column = 9).value.encode('utf8')+'@'+\
                           Company_f.cell(row = row_num,column = 10).value.encode('utf8')+'@'+\
                           Company_f.cell(row = row_num,column = 11).value.encode('utf8'))
    '''
    company_ID_List.append((Company_f.cell(row = row_num,column = 1).value,\
                            str(Company_f.cell(row = row_num,column = 9).value),\
                           str(Company_f.cell(row = row_num,column = 10).value),\
                           str(Company_f.cell(row = row_num,column = 11).value)))'''
    row_num += 1

myfile = open("CompanyList.txt", 'w')

for i in company_ID_List:

    print >> myfile, i

myfile.close()