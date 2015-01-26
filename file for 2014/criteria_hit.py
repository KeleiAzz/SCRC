__author__ = 'keleigong'


import openpyxl as ox


rating_model = ox.load_workbook('2013 criteria hit.xlsx')
model_orignal = rating_model.get_sheet_by_name('General information')

model_extract = ox.Workbook()
rating_des = model_extract.create_sheet(0,'test')

target1 = model_orignal.range('F13:U189')
target2 = model_orignal.range('X13:AM189')
target3 = model_orignal.range('AP13:BE189')
target4 = model_orignal.range('BH13:BT189')
target5 = model_orignal.range('X13:AM189')
target6 = model_orignal.range('X13:AM189')


'''
(a,b,c1,d1)
(a,b2,c2,d)

(a4,b4,c,d)


(2,2,3,4)
(9,10,3,4)
(1,5,3,4)
(1,2,6,7)

(1,2, , )
( ,2,3, )
(1, ,3,4)


(1,2,3,4))


(1,2, , )
(1,2,3, )
(1,2, ,4)\
(1, ,3,4)

(1,2,3,4)

(1,8,3,4)
(6,2,5,4)
(7,9,3,4)

(1,2,6,7)
(1, ,3,4)
(1,2, ,4)
( ,2,3,4)

(1,2,6,7)
(1, ,3, )
(1,2,5,4)
( ,2,3, )
( ,2, ,4)
( , ,3,4)

(1,2,7,8)
(1,2,6,4)
(5,2,3,4)
(9,0,3,4)'''