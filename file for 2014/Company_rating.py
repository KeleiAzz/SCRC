__author__ = 'keleigong'
#coding=gbk

import openpyxl as ox
import datetime
Company = ox.load_workbook('Final Ratings.xlsx')  #储存rating by category的数据,将数据格式改编为可以直接导入navicat的格式.
rating = Company.get_sheet_by_name('Sheet1')  #存有rating数据
Company_list = Company.get_sheet_by_name('list')#存有company name 和ID的对应

Company_info = Company_list['A2':'C1247']
Company_info_rating = rating['B2':'J1554']

save_for_navi = ox.Workbook()  #存为navicat可以导入的格式,新建一个workbook
company_category = save_for_navi.create_sheet(0, 'Sheet1') #在新建的workbook里新建sheet

list = []  #储存name 和 list的对应
list2 = [] #

for rows in Company_info:
    list.append(rows[0].value)
    list.append(rows[1].value)
    list.append(rows[2].value)

for rows in Company_info_rating:
    if rows[7].value == u'N/A':
        LHR = 0
    else:
        LHR = rows[7].value

    if rows[8].value == u'N/A':
        SUS = 0
    else:
        SUS = rows[7].value

    name_without_comma = rows[0].value.replace(',', '')
    if name_without_comma in list:
        ID = list[list.index(name_without_comma) - 1]
    elif rows[1].value in list:
        ID = list[list.index(rows[1].value) - 2]
    list2.append([ID,name_without_comma, rows[1].value, rows[2].value, rows[3].value, rows[4].value, \
                  rows[5].value, rows[6].value, LHR, SUS])

'''for i in range(len(list2)):
    if list2[i][0] in list:
        rating.cell(row=i + 2, column=11).value = list[list.index(list2[i][0]) - 1]
    elif list2[i][1] in list:
        rating.cell(row=i + 2, column=11).value = list[list.index(list2[i][1]) - 2]'''

start = 1
#tmp = rating['B2':'K1554']

for rows in list2:
    for i in range(6):
        effecitive_date = datetime.date(rows[3],1,1)
        expired_date = datetime.date(rows[3],12,31)
        company_category.cell(row=start, column=1).value = rows[0] #id
        company_category.cell(row=start, column=2).value = rows[1] #name
        company_category.cell(row=start, column=3).value = rows[2] #ticker
        company_category.cell(row=start, column=4).value = effecitive_date #date
        company_category.cell(row=start, column=5).value = rows[4 + i] #rating
        company_category.cell(row=start, column=6).value = i + 1 #category
        company_category.cell(row=start, column=7).value = expired_date
        start += 1
#save_for_navi.save('rating_by_category.xlsx')
# Company.save('list_new.xlsx')

#for i in range(len(list2)):
#    list2[i][0].replace(',','')