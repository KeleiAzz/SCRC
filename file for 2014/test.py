__author__ = 'keleigong'

from openpyxl import *
wb = Workbook()

ws = wb.active

ws1 = wb.create_sheet()

ws2= wb.create_sheet(0)

ws2.cell('A4').value = 5

wb.save('test111.xlsx')