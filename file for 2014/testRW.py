__author__ = 'keleigong'
import openpyxl as ox

test = ox.load_workbook('/Users/keleigong/Dropbox/Python/Project-SCRC/test1.xlsx')

target = test.get_sheet_by_name('Sheet2')
src = test.get_sheet_by_name('New 1')

cList = src['A1':'C1246']
name_id = {}
ticker_id = {}
ticker_name = {}
for row in cList:
    name_id[row[1].value]=row[0].value
    ticker_id[row[2].value]=row[0].value
    ticker_name[row[2].value] = row[1].value
for row in range(13,1190,1):
    temp = target.cell('B%s'%(row)).value.replace(',','')
    temp2 = target.cell('C%s'%(row)).value
    if temp in name_id.keys():
        target.cell('A%s'%(row)).value = name_id[temp]
    elif temp2 in ticker_id.keys() and temp2 != u'#N/A' and type(temp2) != type(target.cell('A372').value):
        target.cell('A%s'%(row)).value = ticker_id[temp2]
        target.cell('B%s'%(row)).value = ticker_name[temp2]



test.save('test2.xlsx')