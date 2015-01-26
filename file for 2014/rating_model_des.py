__author__ = 'keleigong'


import openpyxl as ox


rating_model = ox.load_workbook('2013 criteria hit.xlsx')
model_orignal = rating_model.get_sheet_by_name('General information')

model_extract = ox.Workbook()
rating_des = model_extract.create_sheet(0,'test')


target = model_orignal.range('EO1:GF12')

cell_range = model_orignal['EO1':'GF12']

temp =[]

for i in range(len(cell_range[0])):
    tt = []
    for j in range(len(cell_range)):
        tt.append(cell_range[j][i].value)

    temp.append(tt)
row_id = 1
for t in temp:
    for i in range(len(t)-2):
        if t[i] != None :
            rating_des.cell(row = row_id, column= 1).value = t[-1]
            rating_des.cell(row = row_id, column = 2).value = t[i]
            row_id += 1

model_extract.save('rating_model.xlsx')