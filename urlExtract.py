__author__ = 'keleigong'
#-*-coding=utf-8-*-
import urllib2
import urllib
import json
import openpyxl as ox
#seachstr = 'apple sustainability'
seachstr2 = 'apple supplier'
seachstr3 = 'apple supplier labor'

# companyList = ['apple', 'nike', 'ibm', 'general mills', 'electronic arts']
companyList = ['general mills']
# keyword = ['supplier', 'Supplier code of conduct', 'Supplier expectation', 'Vendor code of conduct',
           # 'Vendor expectation', 'Supplier guideline']

keyword = ['Ariba',
           'buying channel',
           'Category Management',
           'Contract Management',
           'cross-functional team',
           'Diversity',
           'ERP',
           'Health Safety Security Environment',
           'Offshore Supplier',
           'procurement allocation',
           'Risk management',
           'Service Level Agreement',
           'Sourcing Process',
           'Strategic Sourcing',
           'Supplier assessment',
           'Supplier collaboration',
           'Supplier development plan',
           'Supplier feedback mechanism',
           'Supplier list',
           'Supplier Meeting',
           'Supplier portal',
           'Supplier scorecard',
           'Supplier segmentation',
           'Supplier selection',
           'Supplier suggestion',
           'Supply market',
           'Vendor list']

date = '2013..2014'

# def write_to_exl(res, urlextract):
#     # urlextract = ox.Workbook()
#     sheet = urlextract.create_sheet(0, 'sheet1')
#     f = res.keys()
#     Row = 1
#     for i in f:
#         sheet.cell(row=Row, column=1).value = ('URLs with frequency = %d' % i)
#         Row += 1
#         for url in res[i]:
#             sheet.cell(row=Row, column=1).value = url
#             Row += 1
#     urlextract.save('url_frequency.xlsx')





file=open('result.txt', 'w')
# Row = 1
sheet_no = 0
for company in companyList:
    urlextract = ox.Workbook()
    all_URL = []
    count = []
    sheet = urlextract.create_sheet(sheet_no, company)
    sheet_no += 1
    for key in keyword:
        searchstr = company + ' ' + key + ' ' + date
        for x in xrange(1):
            print "page:%s" % (x+1)
            page = x*9  # page = x

            url = ('https://ajax.googleapis.com/ajax/services/search/web'
                   '?v=2.0&q=%s&rsz=8&start=%s') % (urllib.quote(searchstr), page)
            try:
                request = urllib2.Request(url, None, {'Referer': 'http://www.google.com'})
                response = urllib2.urlopen(request)

            # Process the JSON string.
                results = json.load(response)
                infoaaa = results['responseData']['results']
            except Exception, e:
                print e
            else:
                for minfo in infoaaa:
                    if minfo['url'][4] == 's':
                        url = minfo['url'][8:]
                    else:
                        url = minfo['url'][7:]
                    if url.find('wikipedia.org') < 0 and url.find('linkedin.com') < 0:
                        if url not in all_URL:
                            print url
                            file.write(url+'\n')
                            # sheet.cell(row=Row, column=1).value = company
                            # sheet.cell(row=Row, column=2).value = key
                            # sheet.cell(row=Row, column=3).value = minfo['url']
                            # Row += 1
                            all_URL.append(url)
                            count.append(1)
                        else:
                            count[all_URL.index(url)] += 1
    res = {}
    for i in range(len(count)):
        if count[i] in res.keys():
            res[count[i]].append(all_URL[i])
        else:
            res[count[i]] = [all_URL[i]]
    f = res.keys()
    Row = 1
    for i in f:
        sheet.cell(row=Row, column=1).value = ('URLs with frequency = %d' % i)
        Row += 1
        for url in res[i]:
            # sheet.cell(row=Row, column=1).value = company
            sheet.cell(row=Row, column=2).value = url
            Row += 1

    urlextract.save('url_%s.xlsx' % company)
print 'Program finished'

def write_to_txt(res):
    file = open('res_fq', 'w')
    for i in res.keys():
        file.write('URLs appear %d times \n' % i)
        for url in res[i]:
            file.write('  ' + url + '\n')

    file.close()
file.close()
# urlextract.save('urlextract.xlsx')